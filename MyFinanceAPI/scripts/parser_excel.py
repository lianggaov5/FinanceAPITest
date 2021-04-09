# -*- coding: utf-8 -*-
"""
********************************
@Time     :2021/1/25 17:29
@Author   :John
@Email    :337901080@qq.com
@File     :parser_excel.py
@Software :PyCharm
********************************
"""
from collections import namedtuple

from openpyxl import load_workbook

from MyFinanceAPI.scripts.parser_config import do_config
from MyFinanceAPI.scripts.constants import DATAS_FILE_PATH


class ParserExcel:
    """
    excel的封装
    """

    def __init__(self, filename=DATAS_FILE_PATH, sheet_name=None):
        self.filename = filename
        self.sheet_name = sheet_name
        self.wb = load_workbook(self.filename)
        self.ws = self.wb[self.sheet_name] if self.sheet_name is not None else self.wb.active
        self.sheet_head_tuple = tuple(self.ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    def get_cases(self):
        """
        获取所有的测试数据
        :return:
        """
        sheet_head_tuple = tuple(self.ws.iter_rows(max_row=1, values_only=True))[0]
        Cases = namedtuple("cases", sheet_head_tuple)
        new_list = []
        for row_data in self.ws.iter_rows(min_row=2, values_only=True):
            new_list.append(Cases(*row_data))

        return new_list

    def get_case(self, row):
        """
        获取某一行测试数据
        :param row:
        :return:
        """
        row_data = tuple(self.ws.iter_rows(min_row=row, max_row=row, values_only=True))[0]
        return row_data

    def write_result_to_excel(self, row, actual, result):
        """
        将测试结果写入到excel中
        :param row:
        :param actual:
        :param result:
        :return:
        """

        other_wb = load_workbook(self.filename)
        other_ws = other_wb[self.sheet_name] if self.sheet_name is not None else other_wb.active
        if isinstance(row, int) and 2 <= row <= self.ws.max_row:
            other_ws.cell(row=row, column=do_config("excel", "actual_col")).value = actual
            other_ws.cell(row=row, column=do_config("excel", "result_col")).value = result
            other_wb.save(self.filename)
        else:
            print("传递的行数有误！")


if __name__ == '__main__':
    parser_excel = ParserExcel(DATAS_FILE_PATH, sheet_name="register")
    print(parser_excel.get_cases())
    print(parser_excel.get_case(1))
    parser_excel.write_result_to_excel(2, 6, "Pass")
# 如何来封装一些操作呢？
# 1、明确需求（读取所有的测试用例、读取一条用例、写入结果）
# 2、将一些写死的数据把它提取出来，往往当做实例属性
# 3、如果其他实例方法中需要调用构造方法中的变量，那么需要将其定义为属性（self.属性名 = 属性值）
